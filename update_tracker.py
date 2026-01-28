from openpyxl import load_workbook
import sys

def update_assumptions(file_path):
    try:
        wb = load_workbook(file_path)
        
        # Central Assumptions Sheet
        if 'Assumptions' in wb.sheetnames:
            ws_asm = wb['Assumptions']
            
            # Organic Growth (Row 9)
            organic_rates = [0.04, 0.15, 0.08, 0.15, 0.20]
            for i, rate in enumerate(organic_rates):
                ws_asm.cell(row=9, column=3+i).value = rate
            
            # Paid Growth (Row 11)
            paid_rates = [0.05, 0.10, 0.08, 0.12, 0.10]
            for i, rate in enumerate(paid_rates):
                ws_asm.cell(row=11, column=3+i).value = rate
            
            # AOV Growth (Row 22)
            aov_rates = [0.005, 0.01, 0.005, 0.01, 0.01]
            for i, rate in enumerate(aov_rates):
                ws_asm.cell(row=22, column=3+i).value = rate
                
            # CPC Trend (Row 32)
            cpc_trends = [0.005, 0.005, 0.005, 0.005, 0.005]
            for i, trend in enumerate(cpc_trends):
                ws_asm.cell(row=32, column=3+i).value = trend

            # Paid Spend Growth (Row 30) - Also stress testing
            spend_growth = [0.05, 0.10, 0.08, 0.12, 0.10]
            for i, growth in enumerate(spend_growth):
                ws_asm.cell(row=30, column=3+i).value = growth

        # Individual Site Sheets
        sites = ['GTSE.co.uk', 'GTSEGroup.com', 'SSH UK', 'SSH USA', 'Epha.com']
        
        for i, site in enumerate(sites):
            if site in wb.sheetnames:
                ws = wb[site]
                # Updating local blocks (Row 9=Organic, 11=Paid, 16=AOV, 20=Spend)
                ws.cell(row=9, column=3).value = organic_rates[i]
                ws.cell(row=11, column=3).value = paid_rates[i]
                ws.cell(row=16, column=3).value = aov_rates[i]
                ws.cell(row=20, column=3).value = spend_growth[i]
                # Note: CPC trend is likely only in the central sheet or hidden formulas
        
        wb.save(file_path)
        print(f"Successfully updated central and local assumptions in {file_path}")

    except Exception as e:
        print(f"Error updating Excel file: {e}")
        sys.exit(1)

if __name__ == "__main__":
    update_assumptions(r'c:\AI_Project\GTSE\Finance Demo\Ecommerce_Budget_Tracker.xlsx')
